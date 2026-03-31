# -------- SimpliScreen — Company Services Classifier --------
import streamlit as st
import pandas as pd
import re
import os
import time
import random
import json
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import google.genai as genai
from urllib.parse import urljoin, urlparse, quote_plus
from collections import deque
from bs4 import BeautifulSoup
import httpx
import trafilatura
from io import BytesIO

try:
    import openpyxl
    from openpyxl.utils import coordinate_to_tuple
except ImportError:
    openpyxl = None

# ---------------- Auth ----------------
def _login_screen():
    st.set_page_config(page_title="SimpliScreen", page_icon="🔎", layout="centered")
    st.title("🔎 SimpliScreen")
    st.subheader("Company Services Classifier")
    st.markdown("---")
    st.markdown("### Log in")
    pwd = st.text_input("Password", type="password", placeholder="Enter your access password")
    if st.button("Log in", use_container_width=True):
        customers = st.secrets.get("customers", {})
        for name, customer_pwd in customers.items():
            if pwd == customer_pwd:
                st.session_state["authenticated"] = True
                st.session_state["customer_name"] = name
                st.rerun()
        st.error("Incorrect password. Please contact support if you believe this is an error.")
    st.markdown("---")
    st.caption("Access is subscription-based. Contact us to get started.")

if not st.session_state.get("authenticated"):
    _login_screen()
    st.stop()

# ---------------- Config ----------------
# API key: read from Streamlit secrets (deployed) or env var (local dev)
API_KEY = st.secrets.get("GOOGLE_API_KEY") or os.environ.get("GOOGLE_API_KEY") or "XXXX"
MODEL_NAME = "gemini-2.5-flash"

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

MAX_PAGES_DEFAULT = 40
REQUEST_TIMEOUT = 20.0
MAX_TEXT_PER_PAGE_CHARS = 15000
TOP_SNIPPETS_PER_CATEGORY = 6
THRESHOLD_DEFAULT = 0.6
MAX_PARALLEL_COMPANIES_DEFAULT = 4

CACHE_DIR = Path(os.path.expanduser("~")) / ".crosstree_cache"
CACHE_TTL_HOURS = 24

CAT_ROW = 9
CAT_START_COL = 6
URL_COL = 5
URL_START_ROW = 10
COMPANY_COL = 2

PRIORITY_PATH_KEYWORDS = [
    "service", "solution", "capability", "what-we-do", "offering",
    "platform", "product", "expertise", "consult", "practice",
    "about", "company", "who-we-are", "work", "approach",
]
NOISE_PATH_KEYWORDS = [
    "blog", "news", "press", "event", "career", "job", "privacy",
    "cookie", "legal", "terms", "login", "signin", "404",
]
FALLBACK_PATHS = [
    "/about", "/services", "/solutions", "/what-we-do", "/capabilities",
    "/our-work", "/company", "/team", "/products", "/platform",
]

_daily_quota_exhausted = False

_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_2) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)",
]

# ---------------- Gemini Init ----------------
try:
    client = genai.Client(api_key=API_KEY)
    client.__del__ = lambda self=None: None
except Exception as e:
    client = None

# ---------------- Cache ----------------
CACHE_DIR.mkdir(parents=True, exist_ok=True)

def _cache_key(url):
    return hashlib.md5(url.encode()).hexdigest()

def cache_load(url):
    path = CACHE_DIR / f"{_cache_key(url)}.json"
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text())
        if datetime.now() - datetime.fromisoformat(data["cached_at"]) > timedelta(hours=CACHE_TTL_HOURS):
            return None
        return data["pages"]
    except Exception:
        return None

def cache_save(url, pages):
    try:
        path = CACHE_DIR / f"{_cache_key(url)}.json"
        path.write_text(json.dumps({"url": url, "cached_at": datetime.now().isoformat(), "pages": pages}))
    except Exception:
        pass

def cache_clear_all():
    for f in CACHE_DIR.glob("*.json"):
        f.unlink(missing_ok=True)

def cache_info():
    files = list(CACHE_DIR.glob("*.json"))
    if not files:
        return 0, None
    dates = []
    for f in files:
        try:
            data = json.loads(f.read_text())
            dates.append(datetime.fromisoformat(data["cached_at"]))
        except Exception:
            pass
    return len(files), min(dates) if dates else None

# ---------------- Crawling ----------------
def same_domain(u1, u2):
    return urlparse(u1).netloc == urlparse(u2).netloc

def clean_url(url):
    parsed = urlparse(url)
    return url if parsed.scheme else "https://" + url

def is_http_url(url):
    return url.startswith("http://") or url.startswith("https://")

def url_priority_score(url):
    path = urlparse(url).path.lower()
    if path in ("", "/"):
        return 100
    score = 0
    for kw in PRIORITY_PATH_KEYWORDS:
        if kw in path:
            score += 10
    for kw in NOISE_PATH_KEYWORDS:
        if kw in path:
            score -= 20
    score -= path.count("/") * 2
    return score

def _fetch_one(url, ua_index=0, verify=True):
    headers = dict(BROWSER_HEADERS)
    headers["User-Agent"] = _USER_AGENTS[ua_index % len(_USER_AGENTS)]
    with httpx.Client(timeout=REQUEST_TIMEOUT, headers=headers,
                      follow_redirects=True, verify=verify) as s:
        r = s.get(url)
        r.raise_for_status()
        return r.text, str(r.url)  # return final URL after redirects

def fetch_html_with_ua_rotation(url):
    """Try multiple UAs. Falls back to SSL-disabled and HTTP scheme on failure."""
    last_err = None
    parsed = urlparse(url)

    # Build candidate URLs: original, then http:// variant
    candidates = [url]
    if parsed.scheme == "https":
        candidates.append("http://" + url[8:])

    for candidate in candidates:
        for i in range(len(_USER_AGENTS)):
            try:
                html, final_url = _fetch_one(candidate, ua_index=i)
                return html, final_url, None
            except httpx.HTTPStatusError as e:
                code = e.response.status_code
                if code in (404, 410):
                    return None, None, f"HTTP {code}"
                last_err = f"HTTP {code}"
                time.sleep(random.uniform(0.5, 1.5))
            except httpx.ConnectError as e:
                err_str = str(e)
                # SSL error — retry without verification
                if "SSL" in err_str or "certificate" in err_str.lower():
                    try:
                        html, final_url = _fetch_one(candidate, ua_index=i, verify=False)
                        return html, final_url, None
                    except Exception as e2:
                        last_err = f"SSL error (verify=False also failed): {e2}"
                else:
                    return None, None, f"ConnectError: {e}"
            except httpx.TimeoutException:
                last_err = "Timeout"
            except Exception as e:
                last_err = str(e)

    return None, None, last_err or "All attempts failed"

def fetch_google_cache(url):
    cache_url = f"https://webcache.googleusercontent.com/search?q=cache:{quote_plus(url)}&hl=en"
    try:
        headers = dict(BROWSER_HEADERS)
        headers["User-Agent"] = _USER_AGENTS[0]
        with httpx.Client(timeout=REQUEST_TIMEOUT, headers=headers, follow_redirects=True) as s:
            r = s.get(cache_url)
            r.raise_for_status()
            return (r.text, None) if len(r.text) > 500 else (None, "Google cache empty")
    except Exception as e:
        return None, str(e)

def fetch_bing_cache(url):
    cache_url = f"https://cc.bingj.com/cache.aspx?url={quote_plus(url)}&q={quote_plus(urlparse(url).netloc)}"
    try:
        headers = dict(BROWSER_HEADERS)
        headers["User-Agent"] = _USER_AGENTS[1]
        with httpx.Client(timeout=REQUEST_TIMEOUT, headers=headers, follow_redirects=True) as s:
            r = s.get(cache_url)
            r.raise_for_status()
            return (r.text, None) if len(r.text) > 500 else (None, "Bing cache empty")
    except Exception as e:
        return None, str(e)

def extract_text(html, url):
    try:
        txt = trafilatura.extract(html, url=url, favor_precision=True)
        if txt and len(txt) > 0:
            return txt.strip()
    except Exception:
        pass
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return re.sub(r"\s+", " ", soup.get_text(" ", strip=True))[:MAX_TEXT_PER_PAGE_CHARS]

def get_links(html, base_url):
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("#", "mailto:", "tel:")):
            continue
        absu = urljoin(base_url, href)
        if is_http_url(absu) and same_domain(absu, base_url):
            links.append(absu.split("#")[0])
    return list(dict.fromkeys(links))

def crawl_site(start_url, max_pages=10):
    start_url = clean_url(start_url)
    cached = cache_load(start_url)
    if cached is not None:
        return cached, None, True

    visited = set()
    pages = []
    errors = []
    base_domain = urlparse(start_url).netloc

    # Priority-sorted queue: (negative_score, url)
    queue = [(-url_priority_score(start_url), start_url)]

    while queue and len(pages) < max_pages:
        queue.sort(key=lambda x: x[0])
        _, url = queue.pop(0)

        if url in visited:
            continue
        visited.add(url)

        html, final_url, err = fetch_html_with_ua_rotation(url)

        if html is None:
            is_homepage = (url == start_url)
            is_block = err and ("HTTP 403" in err or "HTTP 429" in err or "HTTP 401" in err)
            is_gone = err and ("HTTP 404" in err or "HTTP 410" in err)

            if is_homepage and (is_block or is_gone or "ConnectError" in str(err)):
                # Try known subpages as fallback
                base = f"{urlparse(start_url).scheme}://{base_domain}"
                for path in FALLBACK_PATHS:
                    fb = base + path
                    if fb not in visited:
                        queue.append((-url_priority_score(fb), fb))
            errors.append(f"{err}: {url}")
            continue

        # If redirect landed on a different domain, update base for link-following
        if final_url:
            final_domain = urlparse(final_url).netloc
            if final_domain and final_domain != base_domain:
                base_domain = final_domain

        text = extract_text(html, final_url or url)
        if not text.strip():
            continue

        pages.append({"url": final_url or url, "text": text[:MAX_TEXT_PER_PAGE_CHARS]})

        if len(pages) < max_pages:
            for link in get_links(html, final_url or url)[:30]:
                if link not in visited:
                    queue.append((-url_priority_score(link), link))

        time.sleep(random.uniform(0.3, 1.0))

    # Cache fallbacks when live crawl yielded nothing
    if len(pages) == 0:
        targets = [start_url] + [
            f"{urlparse(start_url).scheme}://{urlparse(start_url).netloc}{p}"
            for p in FALLBACK_PATHS[:5]
        ]
        for target in targets:
            if len(pages) >= max(3, max_pages // 4):
                break
            for fetch_fn, label in [(fetch_google_cache, "Google Cache"), (fetch_bing_cache, "Bing Cache")]:
                html, err = fetch_fn(target)
                if html:
                    text = extract_text(html, target)
                    if text.strip():
                        pages.append({"url": f"[{label}] {target}", "text": text[:MAX_TEXT_PER_PAGE_CHARS]})
                        break

    if len(pages) == 0:
        reason = "; ".join(errors[:3]) if errors else "unknown"
        return [], f"All crawl methods failed: {reason}", False

    cache_save(start_url, pages)
    return pages, None, False

# ---------------- Snippets ----------------
def chunk_text(text, max_chars=1200, overlap=150):
    chunks, i = [], 0
    while i < len(text):
        j = min(len(text), i + max_chars)
        chunks.append(text[i:j])
        i = j - overlap
        if i < 0:
            i = 0
        if j == len(text):
            break
    return chunks

def build_snippets(pages):
    snippets = []
    for p in pages:
        for ch in chunk_text(p["text"]):
            if len(ch) >= 200:
                snippets.append({"url": p["url"], "text": ch})
    return snippets[:500]

def rank_snippets_for_category(snippets, category):
    terms = re.findall(r"[a-z0-9]+", category.lower())
    scores = [(sum(sn["text"].lower().count(t) for t in terms) +
               sum(0.5 * sn["text"].lower().count(w)
                   for w in ["service", "solution", "offer", "capability", "platform", "consult"]),
               i)
              for i, sn in enumerate(snippets)]
    scores.sort(reverse=True)
    return [snippets[i] for _, i in scores[:TOP_SNIPPETS_PER_CATEGORY]]

def company_name_from_url(url):
    """Fallback: derive readable name from domain when column B is empty/numeric."""
    try:
        domain = re.sub(r"^www\.", "", urlparse(clean_url(url)).netloc)
        return domain.split(".")[0].replace("-", " ").replace("_", " ").title()
    except Exception:
        return url

# ---------------- Gemini ----------------
def _is_daily_quota_error(err_str):
    signals = ["daily", "per day", "quota exceeded", "has been exhausted",
               "generativelanguage.googleapis.com/default-per-day"]
    return any(s in err_str.lower() for s in signals)

def judge_categories_batch(company_url, categories, snippets_by_cat, user_threshold):
    global _daily_quota_exhausted

    if client is None:
        return {cat: {"has_service": False, "confidence": 0.0, "evidence": [], "notes": "Gemini client not initialized."} for cat in categories}

    parts = [f"""
You are classifying whether the company at {company_url} offers services across multiple categories.
For each category you will receive the strongest evidence snippets.

USER REQUIREMENT:
If you output "has_service": true, "confidence" MUST be >= {user_threshold}.
If your natural confidence is below {user_threshold}, force: "has_service": false, "confidence": 0.0.
If the category is not offered: "has_service": false, "confidence": 0.0.

Return ONLY valid JSON:
{{
  "Category 1": {{"has_service": true/false, "confidence": 0.0-1.0, "evidence": ["quote1", ...], "notes": "short explanation"}},
  "Category 2": {{...}}
}}
"""]
    for cat in categories:
        parts.append(f"\nCATEGORY: \"{cat}\"\nEvidence:")
        for sn in snippets_by_cat.get(cat, []):
            parts.append(f"[{sn['url']}] {sn['text']}\n")

    prompt = "\n".join(parts)

    def call():
        return json.loads(client.models.generate_content(
            model=MODEL_NAME, contents=prompt,
            config=genai.types.GenerateContentConfig(temperature=0.2, response_mime_type="application/json")
        ).text)

    for attempt in range(5):
        try:
            return call()
        except Exception as e:
            err = str(e)
            if "RESOURCE_EXHAUSTED" in err or "429" in err:
                if _is_daily_quota_error(err):
                    _daily_quota_exhausted = True
                    raise RuntimeError(f"DAILY_QUOTA_EXHAUSTED: {err}") from e
                m = re.search(r"retryDelay.*?(\d+)s", err)
                wait = max(int(m.group(1)) + 5 if m else 0, 65 * (2 ** attempt))
                time.sleep(wait)
            else:
                return {cat: {"has_service": False, "confidence": 0.0, "evidence": [], "notes": f"Model error: {e}"} for cat in categories}

    return {cat: {"has_service": False, "confidence": 0.0, "evidence": [], "notes": "Max retries exceeded"} for cat in categories}

# ---------------- Parallel Classify ----------------
def classify_batch_parallel(url_rows, ws, wb, categories, category_positions,
                             max_pages, threshold, n_workers, timing_log, log_box, progress):
    global _daily_quota_exhausted
    total = len(url_rows)
    all_results = []
    low_quality = []

    # Company name lookup — data_only=True means formulas return cached values
    company_names = {}
    for row_idx, url_val in url_rows:
        raw = ws.cell(row=row_idx, column=COMPANY_COL).value
        name = str(raw).strip() if raw is not None else ""
        if not name or name.isdigit():
            name = company_name_from_url(url_val)
        company_names[url_val] = name

    # Pre-populate log rows
    for i, (_, url_val) in enumerate(url_rows):
        timing_log.append({
            "№": i + 1,
            "company": company_names[url_val],
            "url": url_val,
            "status": "⏳ queued",
            "pages": None,
            "source": None,
            "time (s)": None,
        })
    log_box.dataframe(pd.DataFrame(timing_log), use_container_width=True, height=300)

    # Phase 1: Parallel crawl
    crawl_results = {}

    def do_crawl(args):
        i, row_idx, url_val = args
        return i, url_val, crawl_site(url_val, max_pages=max_pages)

    with ThreadPoolExecutor(max_workers=n_workers) as executor:
        futures = {
            executor.submit(do_crawl, (i, row_idx, url_val)): i
            for i, (row_idx, url_val) in enumerate(url_rows)
        }
        for future in as_completed(futures):
            try:
                i, url_val, (pages, err, from_cache) = future.result()
            except Exception as e:
                i = futures[future]
                _, url_val = url_rows[i]
                pages, err, from_cache = [], str(e), False
            crawl_results[url_val] = (pages, err, from_cache)
            timing_log[i]["status"] = "🕸️ crawled" if not err else "❌ crawl error"
            timing_log[i]["pages"] = len(pages)
            timing_log[i]["source"] = "cache" if from_cache else "fresh"
            log_box.dataframe(pd.DataFrame(timing_log), use_container_width=True, height=300)

    # Phase 2: Sequential Gemini calls
    for i, (row_idx, url_val) in enumerate(url_rows):
        if _daily_quota_exhausted:
            for cat in categories:
                ws.cell(row=row_idx, column=category_positions[cat]).value = ""
            timing_log[i]["status"] = "🚫 quota abort"
            log_box.dataframe(pd.DataFrame(timing_log), use_container_width=True, height=300)
            continue

        company_name = company_names[url_val]
        pages, crawl_err, from_cache = crawl_results.get(url_val, ([], "Missing", False))

        if crawl_err:
            for cat in categories:
                ws.cell(row=row_idx, column=category_positions[cat]).value = 0
                all_results.append({"company": company_name, "url": url_val, "category": cat,
                                    "binary": 0, "confidence": 0.0, "notes": f"Crawl error: {crawl_err}"})
            timing_log[i]["status"] = "❌ error"
        else:
            if len(pages) < 3:
                low_quality.append({"company": company_name, "url": url_val, "pages": len(pages)})
                timing_log[i]["status"] = "⚠️ classifying (low data)…"
            else:
                timing_log[i]["status"] = "🤖 classifying…"
            log_box.dataframe(pd.DataFrame(timing_log), use_container_width=True, height=300)

            snippets = build_snippets(pages)
            t0 = time.perf_counter()
            results = {}
            if snippets:
                snippets_by_cat = {cat: rank_snippets_for_category(snippets, cat) for cat in categories}
                try:
                    results = judge_categories_batch(url_val, categories, snippets_by_cat, threshold)
                except RuntimeError as e:
                    if "DAILY_QUOTA_EXHAUSTED" in str(e):
                        timing_log[i]["status"] = "🚫 quota abort"
                        log_box.dataframe(pd.DataFrame(timing_log), use_container_width=True, height=300)
                        progress.progress((i + 1) / total)
                        continue

            elapsed = round(time.perf_counter() - t0, 1)
            for cat in categories:
                r = results.get(cat, {})
                conf = r.get("confidence", 0.0) or 0.0
                value = 1 if conf >= threshold else 0
                ws.cell(row=row_idx, column=category_positions[cat]).value = value
                all_results.append({"company": company_name, "url": url_val, "category": cat,
                                    "binary": value, "confidence": conf,
                                    "has_service": r.get("has_service"), "notes": r.get("notes", ""),
                                    "time_per_company": elapsed})
            timing_log[i]["time (s)"] = elapsed
            timing_log[i]["status"] = "⚠️ done (low data)" if len(pages) < 3 else "✅ done"

        # Interim save after every company
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.session_state["excel_bytes"] = buf.getvalue()

        progress.progress((i + 1) / total)
        log_box.dataframe(pd.DataFrame(timing_log), use_container_width=True, height=300)

    return all_results, low_quality

# ============================= UI =============================
st.set_page_config(page_title="SimpliScreen", page_icon="🔎", layout="wide")

# Header with logout
col_title, col_user = st.columns([5, 1])
with col_title:
    st.title("🔎 SimpliScreen — Company Services Classifier")
with col_user:
    st.markdown(f"<br>", unsafe_allow_html=True)
    customer = st.session_state.get("customer_name", "")
    if st.button(f"Log out ({customer})", use_container_width=True):
        st.session_state.clear()
        st.rerun()

with st.sidebar:
    st.markdown("### ⚙️ Settings")
    max_pages = st.slider("Max pages to crawl per company", 3, 75, MAX_PAGES_DEFAULT, 1)
    st.caption("High-value pages (services, about…) are visited first.")
    threshold = st.slider("Confidence threshold for marking 1 (yes)", 0.0, 1.0, 0.60, 0.01,
                          help="Confidence ≥ threshold → 1, else 0.")
    n_workers = st.slider("Parallel companies to crawl", 1, 8, MAX_PARALLEL_COMPANIES_DEFAULT, 1)
    st.markdown("---")
    st.markdown("### 🗄️ Crawl Cache")
    cache_count, cache_oldest = cache_info()
    if cache_count > 0:
        oldest_str = cache_oldest.strftime("%Y-%m-%d %H:%M") if cache_oldest else "unknown"
        st.caption(f"{cache_count} site(s) cached · oldest: {oldest_str}")
        if st.button("🗑️ Clear cache"):
            cache_clear_all()
            st.success("Cache cleared.")
    else:
        st.caption("No cached sites yet. Sites are cached for 24h after first crawl.")
    st.markdown("---")
    st.markdown("**Excel layout**")
    st.write("- Categories: row 9, starting at column F\n- URLs: column E, starting at row 10\n- Company names: column B")

st.subheader("📂 Upload scorecard Excel")
uploaded_file = st.file_uploader("Upload your scorecard (.xlsx)", type=["xlsx"])

col1, col2 = st.columns(2)
with col1:
    cat_end_cell = st.text_input("Ending cell for categories (last category header)", value="Z9",
                                 help="E.g. Z9 = last category in column Z, row 9.")
with col2:
    url_end_cell = st.text_input("Ending cell for URLs (last company row)", value="E200",
                                 help="E.g. E200 = last URL in row 200, column E.")

if "timing_log" not in st.session_state:
    st.session_state["timing_log"] = []

run = st.button("🚀 Run crawl + classify and write 0/1 into Excel")

if run:
    st.session_state["timing_log"] = []
    st.session_state.pop("results_df", None)
    st.session_state.pop("excel_bytes", None)
    _daily_quota_exhausted = False

    if client is None:
        st.error("Gemini client not initialized. Check your API key in Streamlit secrets.")
    elif uploaded_file is None:
        st.error("Please upload an Excel file first.")
    elif openpyxl is None:
        st.error("openpyxl is not installed.")
    else:
        # Quick API check
        try:
            client.models.generate_content(
                model=MODEL_NAME, contents="Reply with the single word OK.",
                config=genai.types.GenerateContentConfig(temperature=0.0, max_output_tokens=5)
            )
        except Exception as e:
            st.error(f"Gemini API error: {e}")
            st.stop()

        try:
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb[wb.sheetnames[0]]

            # Read categories
            try:
                cat_end_row, cat_end_col = coordinate_to_tuple(cat_end_cell)
            except Exception as e:
                st.error(f"Invalid categories end cell '{cat_end_cell}': {e}")
                st.stop()

            categories, category_positions = [], {}
            for col in range(CAT_START_COL, cat_end_col + 1):
                val = ws.cell(row=CAT_ROW, column=col).value
                if val and str(val).strip():
                    cat_name = str(val).strip()
                    categories.append(cat_name)
                    category_positions[cat_name] = col

            if not categories:
                st.error("No categories found. Check your 'Ending cell for categories'.")
                st.stop()

            # Read URLs
            try:
                url_end_row, _ = coordinate_to_tuple(url_end_cell)
            except Exception as e:
                st.error(f"Invalid URL end cell '{url_end_cell}': {e}")
                st.stop()

            url_rows = []
            for row in range(URL_START_ROW, url_end_row + 1):
                val = ws.cell(row=row, column=URL_COL).value
                if val and str(val).strip():
                    url_rows.append((row, str(val).strip()))

            if not url_rows:
                st.error("No URLs found. Check your 'Ending cell for URLs'.")
                st.stop()

            st.info(f"Found {len(url_rows)} companies · {len(categories)} categories · {n_workers} parallel crawlers")
            progress = st.progress(0)
            timing_log = st.session_state["timing_log"]
            log_box = st.empty()

            all_results, low_quality = classify_batch_parallel(
                url_rows, ws, wb, categories, category_positions,
                max_pages, threshold, n_workers, timing_log, log_box, progress
            )

            if _daily_quota_exhausted:
                st.error("🚫 Daily API quota exhausted. Partial results saved — download below.")
            else:
                st.success(f"✅ Done! Classified {len(url_rows)} companies across {len(categories)} categories.")

            if low_quality:
                names = ", ".join(f"{c['company']} ({c['pages']} pg)" for c in low_quality)
                st.warning(f"⚠️ Low crawl data for: {names}. Results may be less accurate.")

            df_results = pd.DataFrame(all_results)
            if not df_results.empty:
                df_results["binary"] = df_results["binary"].astype(int)
                df_results = df_results.sort_values(["company", "binary", "confidence"],
                                                    ascending=[True, False, False])
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            st.session_state["results_df"] = df_results
            st.session_state["excel_bytes"] = out.getvalue()

        except Exception as e:
            import traceback
            st.error(f"Error: {e}")
            st.code(traceback.format_exc())

# Results + download
if "results_df" in st.session_state and not st.session_state["results_df"].empty:
    st.subheader("📊 Results overview")
    st.dataframe(st.session_state["results_df"], use_container_width=True, height=400)

if "excel_bytes" in st.session_state and st.session_state["excel_bytes"]:
    label = "⬇️ Download partial results Excel" if _daily_quota_exhausted else "⬇️ Download updated Excel"
    st.download_button(label, data=st.session_state["excel_bytes"],
                       file_name="simpliscreen_results.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
